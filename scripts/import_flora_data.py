"""Convert the professor's floristic survey into the JSON the app consumes.

Inputs
  * Floristic list  (.xlsx) - 355 species x 81 families with habit / life form / status
  * Map locations   (.csv)  - GPS fixes recorded per individual plant

Outputs (src/data/generated/)
  * families.json
  * species.json
  * markers.json
  * import-report.json

Run:  python scripts/import_flora_data.py
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "src" / "data" / "generated"

DEFAULT_XLSX = Path.home() / "Downloads" / "Floristic list of Uniflora_2026730 (2).xlsx"
DEFAULT_CSV = Path.home() / "Downloads" / "map_locations2026730 (1).csv"

# The survey is at the University of Sargodha; every fix must fall inside this box.
LAT_DEGREE, LNG_DEGREE = 32, 72

# A fix further than this from the median of the survey cannot be on campus.
CAMPUS_RADIUS_M = 1500

# Field spellings that are unambiguously the same taxon as a listed species.
NAME_ALIASES = {
    "sanchesia": "sanchezia oblonga",
}

PALM_FAMILIES = {"Arecaceae"}

HABIT_TO_TYPE = {
    "tree": "Tree",
    "shrub": "Shrub",
    "subshrub": "Subshrub",
    "succulent shrub": "Succulent",
    "succulent": "Succulent",
    "herb": "Herb",
    "grass": "Grass",
    "sedge": "Sedge",
    "climber": "Climber",
}

TYPE_TO_LAYER = {
    "Tree": "trees",
    "Palm": "trees",
    "Shrub": "shrubs",
    "Subshrub": "shrubs",
    "Climber": "shrubs",
    "Succulent": "herbs",
    "Herb": "herbs",
    "Grass": "herbs",
    "Sedge": "herbs",
}

TYPE_COLORS = {
    "Tree": "#2e6b3a",
    "Palm": "#3f8f5c",
    "Shrub": "#c99a2e",
    "Subshrub": "#b8862b",
    "Climber": "#8163a8",
    "Succulent": "#4f8f77",
    "Herb": "#5a8a2e",
    "Grass": "#8a9a3a",
    "Sedge": "#6f8a4a",
}


# --------------------------------------------------------------------------- #
# text helpers
# --------------------------------------------------------------------------- #

def clean(value) -> str:
    """Normalise whitespace, NBSPs and stray control characters."""
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").replace("\ufeff", "")
    text = unicodedata.normalize("NFC", text)
    return re.sub(r"\s+", " ", text).strip()


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", clean(text))
    text = text.encode("ascii", "ignore").decode("ascii").lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def split_name(raw: str) -> tuple[str, str, str, str]:
    """Return (binomial, genus, epithet, author) from a messy herbarium name."""
    name = clean(raw).replace("?", " ")
    name = re.sub(r"\s+", " ", name).strip()
    tokens = name.split(" ")
    if not tokens:
        return "", "", "", ""

    genus = tokens[0].strip(".,")
    rest = tokens[1:]
    if not rest:
        return genus, genus, "", ""

    # "triquetrumWilld." -> epithet "triquetrum", author starts at "Willd."
    head = rest[0]
    match = re.match(r"^([a-z][a-z-]*)([A-Z(].*)$", head)
    if match:
        epithet = match.group(1)
        author_tokens = [match.group(2)] + rest[1:]
    else:
        epithet = head.strip(".,")
        author_tokens = rest[1:]

    binomial = f"{genus} {epithet}".strip()
    return binomial, genus, epithet, " ".join(author_tokens).strip()


def binomial_key(raw: str) -> str:
    binomial, _, _, _ = split_name(raw)
    key = binomial.lower()
    return NAME_ALIASES.get(key, key)


# --------------------------------------------------------------------------- #
# coordinate parsing / repair
# --------------------------------------------------------------------------- #

class CoordRepair:
    """Fixes the handful of data-entry slips in the GPS column.

    All repairs are conservative: a value is only rewritten when the intended
    reading is unambiguous (a missing space, or a degree that cannot exist for
    this campus). Anything else is reported and dropped.
    """

    def __init__(self, expected_degree: int):
        self.expected = expected_degree
        self.log: Counter = Counter()

    def parse(self, raw: str) -> float | None:
        text = clean(raw)
        if not text:
            self.log["empty"] += 1
            return None

        parts = text.split(" ")

        # "7240 51" -> degrees and minutes ran together
        if len(parts) == 2 and parts[0].startswith(str(self.expected)) and len(parts[0]) > 2:
            merged = parts[0]
            parts = [merged[:2], merged[2:], parts[1]]
            self.log["merged_degree_minute"] += 1
        # "32 427" -> minutes and seconds ran together
        elif len(parts) == 2 and len(parts[1]) >= 3:
            parts = [parts[0], parts[1][:-2], parts[1][-2:]]
            self.log["merged_minute_second"] += 1

        if len(parts) != 3:
            self.log["unparseable"] += 1
            return None

        try:
            deg, minute, second = (float(p) for p in parts)
        except ValueError:
            self.log["non_numeric"] += 1
            return None

        # "74 40 55" / "32 40 57" in the longitude column -> wrong degree typed
        if int(deg) != self.expected:
            self.log[f"degree_{int(deg)}_to_{self.expected}"] += 1
            deg = float(self.expected)

        # seconds recorded as 60 -> roll into the next minute
        if second >= 60:
            minute += second // 60
            second = second % 60
            self.log["second_rollover"] += 1

        if not (0 <= minute < 60 and 0 <= second < 60):
            self.log["out_of_range"] += 1
            return None

        return deg + minute / 60 + second / 3600


# --------------------------------------------------------------------------- #
# geometry: zones from the real point cloud
# --------------------------------------------------------------------------- #

def kmeans(points: list[tuple[float, float]], k: int, iterations: int = 60):
    """Tiny deterministic k-means (no numpy/sklearn dependency)."""
    ordered = sorted(points)
    step = max(1, len(ordered) // k)
    centroids = [ordered[min(i * step, len(ordered) - 1)] for i in range(k)]

    assignment = [0] * len(points)
    for _ in range(iterations):
        moved = False
        for i, (lat, lng) in enumerate(points):
            best, best_dist = 0, float("inf")
            for c, (clat, clng) in enumerate(centroids):
                dist = (lat - clat) ** 2 + (lng - clng) ** 2
                if dist < best_dist:
                    best, best_dist = c, dist
            if assignment[i] != best:
                assignment[i] = best
                moved = True

        for c in range(k):
            members = [points[i] for i, a in enumerate(assignment) if a == c]
            if members:
                centroids[c] = (
                    sum(p[0] for p in members) / len(members),
                    sum(p[1] for p in members) / len(members),
                )
        if not moved:
            break

    return assignment, centroids


def convex_hull(points: list[tuple[float, float]]) -> list[tuple[float, float]]:
    pts = sorted(set(points))
    if len(pts) <= 2:
        return pts

    def cross(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    lower = []
    for p in pts:
        while len(lower) >= 2 and cross(lower[-2], lower[-1], p) <= 0:
            lower.pop()
        lower.append(p)

    upper = []
    for p in reversed(pts):
        while len(upper) >= 2 and cross(upper[-2], upper[-1], p) <= 0:
            upper.pop()
        upper.append(p)

    return lower[:-1] + upper[:-1]


def pad_hull(hull, centroid, factor=1.12):
    return [
        (
            round(centroid[0] + (lat - centroid[0]) * factor, 7),
            round(centroid[1] + (lng - centroid[1]) * factor, 7),
        )
        for lat, lng in hull
    ]


COMPASS_16 = [
    "North", "North-North-East", "North-East", "East-North-East",
    "East", "East-South-East", "South-East", "South-South-East",
    "South", "South-South-West", "South-West", "West-South-West",
    "West", "West-North-West", "North-West", "North-North-West",
]


def compass_name(centroid, center) -> str:
    """Bearing of a zone centroid from the campus centre, as a compass label."""
    dlat = centroid[0] - center[0]
    dlng = (centroid[1] - center[1]) * math.cos(math.radians(center[0]))
    if math.hypot(dlat, dlng) < 0.00015:
        return "Central"
    bearing = (math.degrees(math.atan2(dlng, dlat)) + 360) % 360
    return COMPASS_16[round(bearing / 22.5) % 16]


def haversine_m(a, b) -> float:
    lat1, lng1, lat2, lng2 = map(math.radians, (a[0], a[1], b[0], b[1]))
    dlat, dlng = lat2 - lat1, lng2 - lng1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
    return 2 * 6371000 * math.asin(math.sqrt(h))


def median(values: list[float]) -> float:
    ordered = sorted(values)
    mid = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[mid]
    return (ordered[mid - 1] + ordered[mid]) / 2


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #

def read_species(xlsx: Path):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Sheet1"]

    records = []
    family = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=7, values_only=True):
        _sr, fam, sp, local, status, habit, life = row
        if clean(fam):
            family = clean(fam)
        name = clean(sp)
        if not name or not family:
            continue

        binomial, genus, epithet, author = split_name(name)
        if not binomial:
            continue

        habit_raw = clean(habit)
        habit_key = habit_raw.lower()
        plant_type = HABIT_TO_TYPE.get(habit_key, "Herb")
        if family in PALM_FAMILIES:
            plant_type = "Palm"

        local_names = [n for n in (clean(x) for x in clean(local).split(",")) if n]
        status_clean = clean(status).title() or "Wild"

        records.append(
            {
                "scientificName": binomial,
                "author": author,
                "genus": genus,
                "epithet": epithet,
                "family": family,
                "localNames": local_names,
                "commonName": local_names[0] if local_names else binomial,
                "growthStatus": "Cultivated" if status_clean.startswith("Cult") else "Wild",
                "habit": habit_raw or "Herb",
                "lifeForm": clean(life).title() or "Perennial",
                "type": plant_type,
                "layer": TYPE_TO_LAYER.get(plant_type, "herbs"),
                "badgeColor": TYPE_COLORS.get(plant_type, "#5a8a2e"),
            }
        )

    # unique, stable slugs
    used: Counter = Counter()
    for rec in records:
        base = slugify(rec["scientificName"])
        used[base] += 1
        rec["slug"] = base if used[base] == 1 else f"{base}-{used[base]}"

    return records


def read_markers(csv_path: Path, species_by_binomial, species_by_genus):
    lat_repair = CoordRepair(LAT_DEGREE)
    lng_repair = CoordRepair(LNG_DEGREE)

    with open(csv_path, newline="", encoding="cp1252") as fh:
        rows = list(csv.DictReader(fh))

    resolved = []
    unmatched: Counter = Counter()
    dropped = 0

    for row in rows:
        raw_name = clean(row.get("scientific_name"))
        if not raw_name:
            dropped += 1
            continue

        lat = lat_repair.parse(row.get("latitude", ""))
        lng = lng_repair.parse(row.get("longitude", ""))
        if lat is None or lng is None:
            dropped += 1
            continue

        key = binomial_key(raw_name)
        species = species_by_binomial.get(key)

        if species is None:
            # fall back to genus when it is unambiguous in the floristic list
            genus = key.split(" ")[0].title()
            candidates = species_by_genus.get(genus, [])
            if len(candidates) == 1:
                species = candidates[0]
                unmatched[f"{raw_name}  ->  {species['scientificName']} (genus match)"] += 1
            else:
                unmatched[f"{raw_name}  ->  UNRESOLVED"] += 1
                dropped += 1
                continue

        resolved.append(
            {
                "slug": species["slug"],
                "lat": lat,
                "lng": lng,
                "layer": species["layer"],
                "raw": (clean(row.get("latitude")), clean(row.get("longitude"))),
                "name": raw_name,
            }
        )

    # Reject fixes that cannot be on campus. These are transposition typos in the
    # source (e.g. "72 54 53" among a run of "72 40 53"); they are reported rather
    # than guessed at, because inventing coordinates for survey data is not safe.
    hub = (median([m["lat"] for m in resolved]), median([m["lng"] for m in resolved]))
    outliers: Counter = Counter()
    on_campus = []
    for m in resolved:
        distance = haversine_m((m["lat"], m["lng"]), hub)
        if distance > CAMPUS_RADIUS_M:
            outliers[f"lat={m['raw'][0]!r} lng={m['raw'][1]!r} ({distance / 1000:.1f} km off) — {m['name']}"] += 1
            dropped += 1
        else:
            on_campus.append({k: m[k] for k in ("slug", "lat", "lng", "layer")})

    return on_campus, unmatched, dropped, lat_repair.log, lng_repair.log, outliers


def jitter(index: int, total: int) -> tuple[float, float]:
    """Spread co-located records around their shared fix.

    Source precision is a whole arc-second (~31 m), so a few metres of spread
    keeps every individual clickable without implying accuracy we do not have.
    """
    if total <= 1:
        return 0.0, 0.0
    golden = 2.399963229728653  # radians
    radius_m = 3.2 * math.sqrt(index + 1)
    angle = index * golden
    dlat = (radius_m * math.cos(angle)) / 111320
    dlng = (radius_m * math.sin(angle)) / (111320 * math.cos(math.radians(LAT_DEGREE)))
    return dlat, dlng


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    parser.add_argument("--zones", type=int, default=6)
    args = parser.parse_args()

    for path in (args.xlsx, args.csv):
        if not path.exists():
            raise SystemExit(f"Source file not found: {path}")

    species = read_species(args.xlsx)
    by_binomial = {s["scientificName"].lower(): s for s in species}
    by_genus = defaultdict(list)
    for s in species:
        by_genus[s["genus"]].append(s)

    markers, unmatched, dropped, lat_log, lng_log, outliers = read_markers(
        args.csv, by_binomial, by_genus
    )

    # ----- zones from the actual point cloud -------------------------------- #
    distinct = sorted({(round(m["lat"], 6), round(m["lng"], 6)) for m in markers})
    assignment, centroids = kmeans(distinct, args.zones)

    center = (
        sum(p[0] for p in distinct) / len(distinct),
        sum(p[1] for p in distinct) / len(distinct),
    )

    clusters = defaultdict(list)
    for point, cluster in zip(distinct, assignment):
        clusters[cluster].append(point)

    # order zones north-west -> south-east so the letters read sensibly
    order = sorted(clusters, key=lambda c: (-centroids[c][0], centroids[c][1]))
    zones, point_to_zone = [], {}
    palette = ["#2e6b3a", "#3a7d47", "#4f8f43", "#5a9e52", "#6bab5f", "#7dbf6b", "#8fce7c", "#a1dd8a"]

    used_labels: Counter = Counter()
    for rank, cluster in enumerate(order):
        letter = chr(ord("A") + rank)
        zone_id = f"zone-{letter.lower()}"
        points = clusters[cluster]
        centroid = centroids[cluster]
        hull = convex_hull(points)
        for point in points:
            point_to_zone[point] = zone_id

        label = compass_name(centroid, center)
        used_labels[label] += 1
        if used_labels[label] > 1:
            label = f"{label} {used_labels[label]}"
        short = "Central Campus" if label == "Central" else f"{label} Campus"

        zones.append(
            {
                "id": zone_id,
                "name": f"Zone {letter} — {short}",
                "shortName": short,
                "color": palette[rank % len(palette)],
                "center": [round(centroid[0], 7), round(centroid[1], 7)],
                "polygon": [list(p) for p in pad_hull(hull, centroid)],
                "plantCount": 0,
            }
        )

    zone_index = {z["id"]: z for z in zones}

    # ----- final markers: stable ids, jitter, zone assignment ---------------- #
    grouped = defaultdict(list)
    for m in markers:
        grouped[(round(m["lat"], 6), round(m["lng"], 6))].append(m)

    per_zone_counter: Counter = Counter()
    final_markers = []

    for point in sorted(grouped):
        members = grouped[point]
        zone_id = point_to_zone.get(point, zones[0]["id"])
        zone_letter = zone_id.split("-")[1].upper()

        for i, m in enumerate(sorted(members, key=lambda x: x["slug"])):
            dlat, dlng = jitter(i, len(members))
            per_zone_counter[zone_letter] += 1
            final_markers.append(
                {
                    "id": f"{zone_letter}-{per_zone_counter[zone_letter]:04d}",
                    "slug": m["slug"],
                    "lat": round(point[0] + dlat, 7),
                    "lng": round(point[1] + dlng, 7),
                    "zoneId": zone_id,
                    "layer": m["layer"],
                }
            )
            zone_index[zone_id]["plantCount"] += 1

    # ----- roll marker counts back into species ----------------------------- #
    occurrences = Counter(m["slug"] for m in final_markers)
    zones_by_species = defaultdict(set)
    for m in final_markers:
        zones_by_species[m["slug"]].add(m["zoneId"])

    for s in species:
        s["occurrences"] = occurrences.get(s["slug"], 0)
        s["zones"] = sorted(zones_by_species.get(s["slug"], []))
        s.pop("epithet", None)

    # ----- families ---------------------------------------------------------- #
    families = []
    for name in sorted({s["family"] for s in species}):
        members = [s for s in species if s["family"] == name]
        families.append(
            {
                "slug": slugify(name),
                "name": name,
                "letter": name[0].upper(),
                "genera": sorted({s["genus"] for s in members}),
                "speciesCount": len(members),
                "occurrences": sum(s["occurrences"] for s in members),
                "habits": [h for h, _ in Counter(s["habit"] for s in members).most_common()],
                "cultivated": sum(1 for s in members if s["growthStatus"] == "Cultivated"),
                "wild": sum(1 for s in members if s["growthStatus"] == "Wild"),
            }
        )

    lats = [m["lat"] for m in final_markers]
    lngs = [m["lng"] for m in final_markers]
    bounds = [
        [round(min(lats) - 0.0006, 6), round(min(lngs) - 0.0006, 6)],
        [round(max(lats) + 0.0006, 6), round(max(lngs) + 0.0006, 6)],
    ]

    campus = {
        "center": [round((min(lats) + max(lats)) / 2, 7), round((min(lngs) + max(lngs)) / 2, 7)],
        "bounds": bounds,
        "zoom": 17,
        "zones": zones,
    }

    report = {
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "sources": {"floristicList": args.xlsx.name, "mapLocations": args.csv.name},
        "counts": {
            "families": len(families),
            "species": len(species),
            "markers": len(final_markers),
            "distinctFixes": len(distinct),
            "droppedRows": dropped,
            "speciesWithoutMarkers": sum(1 for s in species if s["occurrences"] == 0),
        },
        "coordinateRepairs": {"latitude": dict(lat_log), "longitude": dict(lng_log)},
        "rejectedOffCampusFixes": dict(outliers),
        "nameResolution": dict(unmatched),
    }

    # Markers ship to the browser, so store them as index-encoded integer offsets
    # from the south-west corner instead of 3k repeated objects (~324 KB -> ~70 KB).
    slug_list = sorted({m["slug"] for m in final_markers})
    slug_index = {s: i for i, s in enumerate(slug_list)}
    zone_list = [z["id"] for z in zones]
    zone_lookup = {z: i for i, z in enumerate(zone_list)}
    layer_list = ["trees", "shrubs", "herbs"]
    layer_lookup = {l: i for i, l in enumerate(layer_list)}

    origin = bounds[0]
    packed = [
        [
            slug_index[m["slug"]],
            round((m["lat"] - origin[0]) * 1e7),
            round((m["lng"] - origin[1]) * 1e7),
            zone_lookup[m["zoneId"]],
            layer_lookup[m["layer"]],
        ]
        for m in final_markers
    ]

    markers_payload = {
        "origin": origin,
        "scale": 1e-7,
        "slugs": slug_list,
        "zones": zone_list,
        "layers": layer_list,
        "markers": packed,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "families.json").write_text(json.dumps(families, indent=1, ensure_ascii=False), "utf-8")
    (OUT_DIR / "species.json").write_text(json.dumps(species, indent=1, ensure_ascii=False), "utf-8")
    (OUT_DIR / "markers.json").write_text(json.dumps(markers_payload, separators=(",", ":"), ensure_ascii=False), "utf-8")
    (OUT_DIR / "campus.json").write_text(json.dumps(campus, indent=1, ensure_ascii=False), "utf-8")
    (OUT_DIR / "import-report.json").write_text(json.dumps(report, indent=1, ensure_ascii=False), "utf-8")

    print(json.dumps(report, indent=1, ensure_ascii=False))
    print()
    print("Zones:")
    for z in zones:
        print(f"  {z['id']}  {z['name']:38s} {z['plantCount']:5d} plants")
    print()
    print(f"Wrote 5 files to {OUT_DIR}")


if __name__ == "__main__":
    main()
